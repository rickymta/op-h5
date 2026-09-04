#!/usr/bin/env python3
"""Xuat toan bo bang cau hinh Excel (server/excel/release/*.xlsx) ra JSON de doc/diff/sua.

VI SAO CO JSON MA SERVER VAN NAP XLSX
    Server la JAR khong co ma nguon: EExcel.load() mo file .xlsx bang POI theo TEN FILE va
    TEN SHEET hardcode trong bytecode. Doi dinh dang server nap la khong the. Vi vay:
        server/excel-src/  (JSON, nguon su that, commit vao git, doc/diff duoc)
            --tools/json-to-excel.py-->  server/excel/release/*.xlsx  (dinh dang server nap)
    Sua cau hinh: sua JSON -> bien dich lai xlsx -> POST :9999/srv/game/cmd/excel/reload.

BO CUC
    server/excel-src/<ten-file-khong-duoi>/_index.json   ten workbook, danh sach sheet theo thu tu
    server/excel-src/<ten-file-khong-duoi>/NN.json       mot sheet (NN = so thu tu, vi ten sheet la
                                                          tieng Trung — khong dung lam ten file de tranh
                                                          hong encoding tren Windows/SSH; ten sheet nam
                                                          trong truong "sheet")
    Moi sheet: {"workbook","sheet","index","rows":[[o1,o2,...],...],"formulas":{"C5":"=A5*2"},
                "merged":["A1:B1"],"dims":{"rows":N,"cols":M}}
      - rows: MOI dong tu dong 1 (dong dau thuong la header) toi dong cuoi co du lieu; moi dong bo
        cac o THIEU o cuoi. O: so (int/float), chuoi, true/false, null (o THIEU - khong co phan tu),
        "" (o CO MAT nhung trong, hoac chuoi rong), {"$error":"#N/A"} (o loi), {"$date":"..."} (o ngay).
        Phan biet null va "" la co chu y: XSSFRowWrap cua server tra gia tri MAC DINH cho o thieu
        nhung tra "" cho o co mat-trong (ca getString lan getInteger), nen phai giu nguyen.
      - Kieu o duoc giu NGUYEN: so la so, chuoi la chuoi. Server doc cot theo ten va doi kieu
        (XSSFRowWrap.getString/getInteger) — "123" va 123 KHAC nhau voi server.
      - O cong thuc: rows chua GIA TRI DA TINH (dieu server nhin thay), formulas giu cong thuc de
        tham khao. Bien dich lai chi ghi gia tri (openpyxl khong ghi duoc cache cua cong thuc).

  python tools/excel-to-json.py                       # tat ca .xlsx trong server/excel/release
  python tools/excel-to-json.py hero.xlsx item.xlsx   # mot vai file
  python tools/excel-to-json.py --src DIR --out DIR

Bo qua: file %3F (rac), .xlsm (macro, khong duoc bytecode tham chieu), file khong phai .xlsx.
"""
import argparse
import datetime as dt
import glob
import io
import json
import os
import sys
import warnings

import openpyxl
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'server', 'excel', 'release')
OUT = os.path.join(ROOT, 'server', 'excel-src')


def cell_json(c):
    v = c.value
    if v is None:
        return None
    t = c.data_type
    if t == 'e':
        return {'$error': str(v)}
    if isinstance(v, dt.datetime):
        return {'$date': v.isoformat()}
    if isinstance(v, dt.date):
        return {'$date': v.isoformat()}
    if isinstance(v, dt.time):
        return {'$time': v.isoformat()}
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v)


def export_workbook(path, out_root):
    name = os.path.basename(path)
    stem = os.path.splitext(name)[0]
    wb_v = openpyxl.load_workbook(path, data_only=True)      # gia tri (cache cua cong thuc)
    wb_f = openpyxl.load_workbook(path, data_only=False)     # cong thuc
    d = os.path.join(out_root, stem)
    os.makedirs(d, exist_ok=True)
    for old in glob.glob(os.path.join(d, '*.json')):
        os.remove(old)
    index = {'workbook': name, 'sheets': []}
    stats = {'sheets': 0, 'cells': 0, 'formulas': 0, 'formula_no_cache': 0}
    for i, ws in enumerate(wb_v.worksheets, 1):
        wsf = wb_f[ws.title]
        max_r, max_c = ws.max_row, ws.max_column
        rows = []
        last_nonempty = 0
        # ws._cells chi chua o CO PHAN TU trong XML (ke ca <c r="D2" s="1"/> khong gia tri). Khong dung
        # ws.iter_rows()/ws['D2'] o day vi chung TAO o moi khi truy cap. Server phan biet ro:
        # XSSFRowWrap.getCell -> null (o thieu) -> tra mac dinh; o co mat nhung trong -> "" (ke ca getInteger).
        cells = ws._cells
        for r_i in range(1, max_r + 1):
            vals = []
            for c_i in range(1, max_c + 1):
                c = cells.get((r_i, c_i))
                if c is None:
                    vals.append(None)
                elif c.value is None:
                    vals.append('')          # co mat nhung trong -> "" (bien dich lai thanh <c r=".."/>)
                else:
                    vals.append(cell_json(c))
            while vals and vals[-1] is None:
                vals.pop()
            rows.append(vals)
            if vals:
                last_nonempty = len(rows)
        rows = rows[:last_nonempty]
        stats['cells'] += sum(1 for row in rows for v in row if v is not None)
        formulas = {}
        for r in wsf.iter_rows(min_row=1, max_row=min(max_r, wsf.max_row), max_col=max_c):
            for c in r:
                if c.data_type == 'f':   # chi o cong thuc that; chuoi '=...' (t="s") van la chuoi
                    ref = f'{get_column_letter(c.column)}{c.row}'
                    formulas[ref] = str(c.value)
                    if ws[ref].value is None:
                        stats['formula_no_cache'] += 1
        stats['formulas'] += len(formulas)
        merged = sorted(str(m) for m in ws.merged_cells.ranges)
        doc = {'workbook': name, 'sheet': ws.title, 'index': i, 'rows': rows}
        if formulas:
            doc['formulas'] = formulas
        if merged:
            doc['merged'] = merged
        doc['dims'] = {'rows': max_r, 'cols': max_c}
        fn = f'{i:02d}.json'
        with io.open(os.path.join(d, fn), 'w', encoding='utf-8', newline='\n') as f:
            json.dump(doc, f, ensure_ascii=False, separators=(',', ':'), indent=None)
            f.write('\n')
        index['sheets'].append({'file': fn, 'name': ws.title, 'rows': len(rows), 'cols': max_c,
                                'formulas': len(formulas)})
        stats['sheets'] += 1
    with io.open(os.path.join(d, '_index.json'), 'w', encoding='utf-8', newline='\n') as f:
        json.dump(index, f, ensure_ascii=False, indent=1)
        f.write('\n')
    return stats


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('files', nargs='*')
    ap.add_argument('--src', default=SRC)
    ap.add_argument('--out', default=OUT)
    a = ap.parse_args()
    if a.files:
        files = [os.path.join(a.src, f) if not os.path.isabs(f) else f for f in a.files]
    else:
        files = sorted(glob.glob(os.path.join(a.src, '*.xlsx')))
    files = [f for f in files if '%3F' not in os.path.basename(f) and not os.path.basename(f).startswith('~$')]
    os.makedirs(a.out, exist_ok=True)
    tot = {'files': 0, 'sheets': 0, 'cells': 0, 'formulas': 0, 'formula_no_cache': 0}
    manifest = {}
    for f in files:
        try:
            st = export_workbook(f, a.out)
        except Exception as e:  # noqa: BLE001
            print(f'LOI {os.path.basename(f)}: {e}', file=sys.stderr)
            continue
        manifest[os.path.basename(f)] = st
        tot['files'] += 1
        for k in ('sheets', 'cells', 'formulas', 'formula_no_cache'):
            tot[k] += st[k]
        print(f"  {os.path.basename(f)}: {st['sheets']} sheet, {st['cells']} o, {st['formulas']} cong thuc")
    with io.open(os.path.join(a.out, '_manifest.json'), 'w', encoding='utf-8', newline='\n') as f:
        json.dump({'source': os.path.relpath(a.src, ROOT).replace('\\', '/'), 'workbooks': manifest},
                  f, ensure_ascii=False, indent=1)
        f.write('\n')
    print(f"xong: {tot['files']} workbook, {tot['sheets']} sheet, {tot['cells']} o, "
          f"{tot['formulas']} cong thuc ({tot['formula_no_cache']} khong co gia tri cache) -> {os.path.relpath(a.out, ROOT)}")


if __name__ == '__main__':
    sys.exit(main())
