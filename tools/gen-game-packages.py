#!/usr/bin/env python3
"""Sinh SQL seed cho bang platform.game_packages tu bang gia cua tang PHP cu.

Nguon:
  website/game/api/id.txt                 payId;giaXu   (1984 dong) — bang gia ma api.php/apisv.php
                                          doi chieu truoc khi tru xu; day la NGUON SU THAT ve gia.
  server/excel/release/recharge-item.xlsx sheet 充值项: ID, *名称, 额度, ... — ten goi (tieng Trung,
                                          chua co ban Viet) va de doi chieu gia.

item_tid = payId: gmhanglong/gm/webshop.php ghi thang payId vao tcg.pay_approval.item_tid,
nen goi nao co trong id.txt la goi ma game phat duoc.

Ket qua: docker/platform-seed/game_packages.<game>.sql — INSERT ... ON DUPLICATE KEY UPDATE,
chay lai bao nhieu lan cung duoc (platform-seed.sh nap file nay sau khi migration xong).
`name`/`status` KHONG bi ghi de khi da co, de sua ten tieng Viet trong DB khong bi mat.

  python tools/gen-game-packages.py                 # ghi file
  python tools/gen-game-packages.py --check         # chi doi chieu, khong ghi
"""
import argparse
import io
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ID_TXT = os.path.join(ROOT, "website", "game", "api", "id.txt")
XLSX = os.path.join(ROOT, "server", "excel", "release", "recharge-item.xlsx")
SHEET = "充值项"
OUT_DIR = os.path.join(ROOT, "docker", "platform-seed")


def read_id_txt(path):
    prices = {}
    order = []
    with io.open(path, encoding="utf-8") as f:
        for ln, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            parts = line.split(";")
            if len(parts) < 2:
                print(f"id.txt dong {ln}: bo qua '{line}'", file=sys.stderr)
                continue
            pid, price = parts[0].strip(), parts[1].strip()
            if not pid.isdigit() or not price.isdigit():
                print(f"id.txt dong {ln}: khong phai so '{line}'", file=sys.stderr)
                continue
            if pid in prices:
                print(f"id.txt dong {ln}: payId {pid} lap, giu dong dau", file=sys.stderr)
                continue
            prices[pid] = int(price)
            order.append(pid)
    return prices, order


def read_excel(path):
    try:
        import openpyxl
    except ImportError:
        print("thieu openpyxl (pip install openpyxl) — ten goi se la 'Goi <id>'", file=sys.stderr)
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"khong thay sheet {SHEET} trong {path}", file=sys.stderr)
        return {}
    ws = wb[SHEET]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    col = {h: i for i, h in enumerate(header)}
    c_id, c_name, c_amt = col.get("ID"), col.get("*名称"), col.get("额度")
    if c_id is None or c_name is None:
        print(f"header khong nhu mong doi: {header}", file=sys.stderr)
        return {}
    items = {}
    for r in rows:
        if r is None or r[c_id] is None:
            continue
        try:
            pid = str(int(r[c_id]))
        except (TypeError, ValueError):
            continue
        name = str(r[c_name]).strip() if r[c_name] is not None else ""
        amt = r[c_amt] if c_amt is not None else None
        items[pid] = (name, int(amt) if isinstance(amt, (int, float)) else None)
    return items


def q(s):
    return s.replace("\\", "\\\\").replace("'", "''")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--game", default="haitac")
    ap.add_argument("--check", action="store_true")
    a = ap.parse_args()

    prices, order = read_id_txt(ID_TXT)
    excel = read_excel(XLSX)

    mismatch = [(p, prices[p], excel[p][1]) for p in order if p in excel and excel[p][1] not in (None, prices[p])]
    no_name = [p for p in order if p not in excel or not excel[p][0]]
    only_excel = [p for p in excel if p not in prices]
    print(f"id.txt: {len(order)} goi; excel: {len(excel)} dong; "
          f"lech gia: {len(mismatch)}; khong co ten: {len(no_name)}; chi co trong excel: {len(only_excel)}")
    for p, a_, b in mismatch[:10]:
        print(f"  lech gia {p}: id.txt={a_} excel={b} -> dung id.txt")
    if a.check:
        return

    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, f"game_packages.{a.game}.sql")
    lines = [
        "-- SINH TU DONG boi tools/gen-game-packages.py — dung sua tay, sua id.txt/Excel roi chay lai.",
        f"-- {len(order)} goi cho game '{a.game}'. Gia tu website/game/api/id.txt, ten tu recharge-item.xlsx (充值项).",
        "-- Ten dang la tieng Trung (bang goc khong co ban Viet): doi trong DB hoac trang quan tri;",
        "-- chay lai file nay KHONG ghi de name/status.",
        "SET NAMES utf8mb4;",
        "INSERT INTO game_packages (game_code, package_id, name, price_xu, item_tid, item_count, item_name, status, sort_order) VALUES",
    ]
    vals = []
    for i, p in enumerate(order):
        name = excel.get(p, ("", None))[0] or f"Goi {p}"
        vals.append(f"('{q(a.game)}','{p}','{q(name)}',{prices[p]},{int(p)},1,'{q(name)}','active',{i})")
    lines.append(",\n".join(vals))
    lines.append("ON DUPLICATE KEY UPDATE price_xu=VALUES(price_xu), item_tid=VALUES(item_tid), "
                 "item_count=VALUES(item_count), item_name=VALUES(item_name), sort_order=VALUES(sort_order);")
    with io.open(out, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")
    print(f"da ghi {os.path.relpath(out, ROOT)} ({len(order)} goi)")


if __name__ == "__main__":
    main()
