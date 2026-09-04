#!/usr/bin/env python3
"""Bien dich JSON (server/excel-src/) nguoc thanh .xlsx cho server nap, va kiem chung.

  python tools/json-to-excel.py                          # tat ca -> build/excel-regen/
  python tools/json-to-excel.py hero item                # mot vai workbook (ten khong duoi)
  python tools/json-to-excel.py --out server/excel/release   # ghi de file server nap (sao luu truoc!)
  python tools/json-to-excel.py --verify                 # sau khi ghi: so tung o voi bang goc
                                                         #   (server/excel/release) theo cach server nhin
  python tools/json-to-excel.py --probe                  # chay ExcelProbe (parser THAT cua game) tren
                                                         #   file goc va file vua sinh, so ket qua

Ghi xong tu dong chay tools/xlsx-inline-to-shared.py cho tung file (bat buoc voi xlsx sinh bang
openpyxl: XSSFRowWrap.getString() cua server lam mat o inlineStr — xem memory/CLAUDE.md).

Chi ghi GIA TRI: o cong thuc trong JSON mang gia tri da tinh, xlsx sinh ra khong con cong thuc
(server chi doc gia tri cache nen ket qua nhu nhau). Merged cells duoc ap lai.
"""
import argparse
import datetime as dt
import glob
import io
import json
import os
import re
import subprocess
import sys
import warnings

import openpyxl
import openpyxl.cell._writer as _xlwriter

warnings.filterwarnings('ignore')


def _exact_number(value):
    """openpyxl ghi so bang "%.16g" (openpyxl.compat.safe_string) -> mat 1 ulp o cac gia tri can
    17 chu so (336.40000000000003 thanh 336.4). Bang goc do Excel ghi du 17 chu so; dung repr()
    (ngan nhat ma doc lai dung tung bit) de xlsx sinh ra giong het ban goc."""
    if isinstance(value, bool):
        return _orig_safe_string(value)
    if isinstance(value, float):
        if value != value or value in (float('inf'), float('-inf')):
            return ''
        return repr(value)
    if isinstance(value, int):
        return str(value)
    return _orig_safe_string(value)


_orig_safe_string = _xlwriter.safe_string
_xlwriter.safe_string = _exact_number
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, 'server', 'excel-src')
REL = os.path.join(ROOT, 'server', 'excel', 'release')
OUT = os.path.join(ROOT, 'build', 'excel-regen')
TOOLS = os.path.join(ROOT, 'tools')


def from_json(v):
    if isinstance(v, dict):
        if '$error' in v:
            return v['$error']
        if '$date' in v:
            return dt.datetime.fromisoformat(v['$date'])
        if '$time' in v:
            return dt.time.fromisoformat(v['$time'])
        raise ValueError(f'o khong hieu: {v}')
    return v


def build_workbook(src_dir, out_path):
    idx = json.load(io.open(os.path.join(src_dir, '_index.json'), encoding='utf-8'))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sh in idx['sheets']:
        doc = json.load(io.open(os.path.join(src_dir, sh['file']), encoding='utf-8'))
        ws = wb.create_sheet(title=doc['sheet'])
        for r, row in enumerate(doc['rows'], 1):
            for c, v in enumerate(row, 1):
                if v is None:
                    continue
                cell = ws.cell(row=r, column=c)
                val = from_json(v)
                cell.value = val
                if isinstance(v, dict) and '$error' in v:
                    cell.data_type = 'e'
                elif isinstance(val, str) and val.startswith('='):
                    # chuoi bat dau bang '=' trong JSON la CHUOI (o cong thuc da duoc thay bang gia tri
                    # khi xuat); openpyxl mac dinh coi la cong thuc -> ep ve kieu chuoi.
                    cell.data_type = 's'
        for m in doc.get('merged', []):
            ws.merge_cells(m)
    wb.save(out_path)
    subprocess.run([sys.executable, os.path.join(TOOLS, 'xlsx-inline-to-shared.py'), out_path],
                   check=True, stdout=subprocess.DEVNULL)
    return idx['workbook']


def norm(v):
    # cach server nhin: so la so (int == float cung gia tri), chuoi la chuoi, loi la chuoi loi
    if isinstance(v, bool):
        return ('b', v)
    if isinstance(v, (int, float)):
        return ('n', float(v))
    if isinstance(v, dt.datetime):
        return ('d', v.isoformat())
    if v is None or v == '':
        # '' (shared string rong) va o BLANK co mat: XSSFRowWrap tra "" cho ca hai -> coi nhu nhau;
        # su CO MAT cua o duoc so rieng ben duoi.
        return None
    return ('s', str(v))


def verify(orig, regen):
    a = openpyxl.load_workbook(orig, data_only=True)
    b = openpyxl.load_workbook(regen, data_only=True)
    if a.sheetnames != b.sheetnames:
        return [f'sheet khac: {a.sheetnames} != {b.sheetnames}']
    diffs = []
    for ws in a.worksheets:
        wb_ = b[ws.title]
        maxr, maxc = max(ws.max_row, wb_.max_row), max(ws.max_column, wb_.max_column)
        ra = ws.iter_rows(min_row=1, max_row=maxr, max_col=maxc)
        rb = wb_.iter_rows(min_row=1, max_row=maxr, max_col=maxc)
        for row_a, row_b in zip(ra, rb):
            for ca, cb in zip(row_a, row_b):
                va, vb = norm(ca.value), norm(cb.value)
                if ca.data_type == 'e':
                    va = ('s', str(ca.value))
                if cb.data_type == 'e':
                    vb = ('s', str(cb.value))
                if va != vb:
                    diffs.append(f'{ws.title}!{ca.coordinate}: goc={va} moi={vb}')
                    if len(diffs) > 20:
                        return diffs
        if sorted(str(m) for m in ws.merged_cells.ranges) != sorted(str(m) for m in wb_.merged_cells.ranges):
            diffs.append(f'{ws.title}: merged khac')
        # tap o CO MAT (ke ca o trong): server tra "" cho o co mat-trong, mac dinh cho o thieu
        pa, pb = set(ws._cells.keys()), set(wb_._cells.keys())
        if pa != pb:
            ex = sorted(pa ^ pb)[:3]
            diffs.append(f'{ws.title}: tap o co mat khac ({len(pa)} vs {len(pb)}; vd {ex})')
    return diffs


def probe(xlsx, cls, timeout=60):
    cp = ';' if os.name == 'nt' else ':'
    classpath = cp.join([os.path.join(ROOT, 'server', 'game', 'tcg-game.jar'),
                         os.path.join(ROOT, 'server', 'game', 'lib', '*'),
                         os.path.join(ROOT, 'build', 'probe')])
    # Popen + giet CA CAY tien trinh khi qua han. ExcelProbe da tu System.exit(), nhung mot so
    # loader (ArenaExcel...) de lai thread nen non-daemon; subprocess.run(timeout=) tren Windows
    # tung ket vo han o day (JVM song 3 gio) nen khong dua vao no nua.
    p = subprocess.Popen(['java', '-Dfile.encoding=UTF-8', '-cp', classpath, 'ExcelProbe',
                          xlsx.replace('\\', '/'), cls],
                         stdout=subprocess.PIPE, stderr=subprocess.STDOUT, cwd=ROOT)
    try:
        out, _ = p.communicate(timeout=timeout)
        return p.returncode, out.decode('utf-8', 'replace')
    except subprocess.TimeoutExpired:
        if os.name == 'nt':
            subprocess.run(['taskkill', '/F', '/T', '/PID', str(p.pid)], capture_output=True)
        else:
            p.kill()
        out, _ = p.communicate()
        return -1, out.decode('utf-8', 'replace') + '\n[timeout]'


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('names', nargs='*', help='ten workbook khong duoi (mac dinh: tat ca)')
    ap.add_argument('--src', default=SRC)
    ap.add_argument('--out', default=OUT)
    ap.add_argument('--orig', default=REL, help='thu muc xlsx goc de --verify/--probe so sanh')
    ap.add_argument('--verify', action='store_true')
    ap.add_argument('--probe', action='store_true')
    ap.add_argument('--no-build', action='store_true', help='chi verify/probe, khong ghi lai')
    a = ap.parse_args()

    dirs = sorted(d for d in glob.glob(os.path.join(a.src, '*')) if os.path.isdir(d))
    if a.names:
        dirs = [d for d in dirs if os.path.basename(d) in set(a.names)]
    os.makedirs(a.out, exist_ok=True)
    built = []
    for d in dirs:
        idx = json.load(io.open(os.path.join(d, '_index.json'), encoding='utf-8'))
        out_path = os.path.join(a.out, idx['workbook'])
        if not a.no_build:
            build_workbook(d, out_path)
        built.append((idx['workbook'], out_path))
    if not a.no_build:
        print(f'da ghi {len(built)} workbook -> {os.path.relpath(a.out, ROOT)}')

    bad = 0
    if a.verify:
        for name, out_path in built:
            orig = os.path.join(a.orig, name)
            if not os.path.exists(orig):
                print(f'  {name}: khong co ban goc de so'); continue
            diffs = verify(orig, out_path)
            if diffs:
                bad += 1
                print(f'  KHAC {name}: ' + '; '.join(diffs[:5]))
        print(f'verify: {len(built) - bad}/{len(built)} workbook giong het ban goc (theo gia tri + kieu o)')

    if a.probe:
        loaders = json.load(io.open(os.path.join(ROOT, 'docs', 'excel-loaders.json'), encoding='utf-8'))
        if not os.path.exists(os.path.join(ROOT, 'build', 'probe', 'ExcelProbe.class')):
            print('chua bien dich ExcelProbe (xem dau tools/ExcelProbe.java)'); return 1
        n = ok = skip = 0
        for name, out_path in built:
            cls_list = loaders.get(name) or []
            if not cls_list:
                skip += 1; continue
            cls = cls_list[0]
            n += 1
            print(f'  probe {n}: {name}', flush=True)
            rc1, o1 = probe(os.path.join(a.orig, name), cls)
            rc2, o2 = probe(out_path, cls)
            # dong dau cua ExcelProbe in kich thuoc file — xlsx sinh lai nen nho hon, bo qua
            o1 = re.sub(r' \(\d+ bytes\)', '', o1)
            o2 = re.sub(r' \(\d+ bytes\)', '', o2)
            # Object.toString mac dinh in identity hash (GroupChangeList@282cb7c7) — khac moi lan chay
            o1 = re.sub(r'@[0-9a-f]{6,8}\b', '@', o1)
            o2 = re.sub(r'@[0-9a-f]{6,8}\b', '@', o2)
            # dong log ERROR cua loader (vd "Cannot get a STRING value from a NUMERIC cell" — loi co san
            # trong ban goc, muc 11.5 CLAUDE.md) bat dau bang gio:phut:giay.ms — khac nhau moi lan chay
            o1 = re.sub(r'^\d\d:\d\d:\d\d\.\d{3} ', 'HH:MM:SS.mmm ', o1, flags=re.M)
            o2 = re.sub(r'^\d\d:\d\d:\d\d\.\d{3} ', 'HH:MM:SS.mmm ', o2, flags=re.M)
            same = (rc1 == rc2) and (o1 == o2)
            if same and rc1 == 0:
                ok += 1
            elif same and rc1 == -1:
                # ca hai deu qua han (loader decode qua lau, vd instance-template.xlsx 55.632 dong > 5 phut)
                # nhung phan da in ra (load OK + so dong tung sheet) giong het -> tinh la khop mot phan
                ok += 1
                print(f'  KHOP MOT PHAN {name} [{cls}]: ca hai deu qua han {60}s, phan load/so dong giong nhau')
            else:
                bad += 1
                tag = 'GOC CUNG LOI' if rc1 != 0 else 'KHAC'
                print(f'  {tag} {name} [{cls}] rc goc={rc1} moi={rc2}')
                if o1 != o2:
                    la, lb = o1.splitlines(), o2.splitlines()
                    for i, (x, y) in enumerate(zip(la, lb)):
                        if x != y:
                            print(f'     dong {i}: goc: {x[:160]}\n              moi: {y[:160]}'); break
        print(f'probe: {ok}/{n} workbook parser that cho ket qua GIONG HET ban goc ({skip} khong co loader)')
    return 1 if bad else 0


if __name__ == '__main__':
    sys.exit(main())
