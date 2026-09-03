#!/usr/bin/env python3
"""
Chen mot sheet moi vao file .xlsx CO SAN o muc XML, khong round-trip qua
openpyxl (openpyxl save se mat cache formula, dinh dang, comment, va ghi chuoi
dang inlineStr ma XSSFRowWrap cua server khong doc duoc).

Chuoi duoc ghi dang shared string (t="s") giong Excel/file goc.

CACH DUNG (nhu thu vien):
    from xlsx_add_sheet import add_sheet
    add_sheet('file.xlsx', 'TenSheet', [['ID','奖励'], [1,'3:100022:1']])

CACH DUNG (CLI) - copy 1 sheet tu workbook khac/cung workbook sang ten moi:
    python tools/xlsx-add-sheet.py <dich.xlsx> <ten_sheet_moi> --from <nguon.xlsx> <ten_sheet_nguon> [--add-cols "cot1=gia_tri,cot2=gia_tri"]

Ghi de tai cho. Neu sheet cung ten da ton tai thi bao loi va khong lam gi.
"""
import io
import re
import sys
import zipfile
from xml.sax.saxutils import escape

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_WS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet'
CT_WS = 'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'
CT_SST = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml'
REL_SST = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings'


def _col(n):  # 0 -> A
    s = ''
    n += 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _unesc(s):
    s = re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))), s)
    s = re.sub(r'&#x([0-9a-fA-F]+);', lambda m: chr(int(m.group(1), 16)), s)
    for a, b in (('&lt;', '<'), ('&gt;', '>'), ('&quot;', '"'), ('&apos;', "'"), ('&amp;', '&')):
        s = s.replace(a, b)
    return s


def add_sheet(path, name, rows):
    zin = zipfile.ZipFile(path)
    files = {n: zin.read(n) for n in zin.namelist()}
    zin.close()

    wb = files['xl/workbook.xml'].decode('utf-8')
    if re.search(r'<sheet [^>]*name="%s"' % re.escape(escape(name, {'"': '&quot;'})), wb):
        raise SystemExit('sheet "%s" da ton tai trong %s' % (name, path))

    # --- shared strings ---
    sst, index = [], {}
    if 'xl/sharedStrings.xml' in files:
        for si in re.findall(r'<si>(.*?)</si>', files['xl/sharedStrings.xml'].decode('utf-8'), re.S):
            t = _unesc(''.join(re.findall(r'<t[^>]*>(.*?)</t>', si, re.S)))
            index.setdefault(t, len(sst)); sst.append(t)

    def sidx(t):
        if t not in index:
            index[t] = len(sst); sst.append(t)
        return index[t]

    # --- sheet xml ---
    body = []
    for r, row in enumerate(rows, 1):
        cells = []
        for c, v in enumerate(row):
            if v is None or v == '':
                continue
            ref = '%s%d' % (_col(c), r)
            if isinstance(v, bool):
                cells.append('<c r="%s" t="b"><v>%d</v></c>' % (ref, int(v)))
            elif isinstance(v, (int, float)):
                cells.append('<c r="%s"><v>%s</v></c>' % (ref, repr(v) if isinstance(v, float) else v))
            else:
                cells.append('<c r="%s" t="s"><v>%d</v></c>' % (ref, sidx(str(v))))
        body.append('<row r="%d">%s</row>' % (r, ''.join(cells)))
    ncols = max((len(r) for r in rows), default=1)
    dim = 'A1:%s%d' % (_col(ncols - 1), max(len(rows), 1))
    sheet_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                 '<worksheet xmlns="%s" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                 '<dimension ref="%s"/><sheetViews><sheetView workbookViewId="0"/></sheetViews>'
                 '<sheetFormatPr defaultRowHeight="15"/><sheetData>%s</sheetData></worksheet>'
                 % (NS, dim, ''.join(body)))

    # --- ten file sheet moi, sheetId, rId ---
    existing = [int(m) for m in re.findall(r'xl/worksheets/sheet(\d+)\.xml', '\n'.join(files))]
    n = max(existing + [0]) + 1
    sheet_file = 'xl/worksheets/sheet%d.xml' % n
    sheet_ids = [int(m) for m in re.findall(r'<sheet [^>]*sheetId="(\d+)"', wb)]
    sheet_id = max(sheet_ids + [0]) + 1
    rels_name = 'xl/_rels/workbook.xml.rels'
    rels = files[rels_name].decode('utf-8')
    rids = [int(m) for m in re.findall(r'Id="rId(\d+)"', rels)]
    rid = max(rids + [0]) + 1

    # workbook.xml: them <sheet> vao cuoi <sheets>
    wb = wb.replace('</sheets>', '<sheet name="%s" sheetId="%d" r:id="rId%d"/></sheets>'
                    % (escape(name, {'"': '&quot;'}), sheet_id, rid), 1)
    files['xl/workbook.xml'] = wb.encode('utf-8')
    # rels
    rels = rels.replace('</Relationships>',
                        '<Relationship Id="rId%d" Type="%s" Target="worksheets/sheet%d.xml"/></Relationships>' % (rid, REL_WS, n), 1)
    if 'xl/sharedStrings.xml' not in files:
        rels = rels.replace('</Relationships>',
                            '<Relationship Id="rId%d" Type="%s" Target="sharedStrings.xml"/></Relationships>' % (rid + 1, REL_SST), 1)
    files[rels_name] = rels.encode('utf-8')
    # content types
    ct = files['[Content_Types].xml'].decode('utf-8')
    ct = ct.replace('</Types>', '<Override PartName="/%s" ContentType="%s"/></Types>' % (sheet_file, CT_WS), 1)
    if 'sharedStrings' not in ct:
        ct = ct.replace('</Types>', '<Override PartName="/xl/sharedStrings.xml" ContentType="%s"/></Types>' % CT_SST, 1)
    files['[Content_Types].xml'] = ct.encode('utf-8')
    # shared strings
    files['xl/sharedStrings.xml'] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="%s" count="%d" uniqueCount="%d">%s</sst>'
        % (NS, len(sst), len(sst), ''.join('<si><t xml:space="preserve">%s</t></si>' % escape(t) for t in sst))).encode('utf-8')
    files[sheet_file] = sheet_xml.encode('utf-8')

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for k, v in files.items():
            z.writestr(k, v)
    with open(path, 'wb') as fh:
        fh.write(buf.getvalue())
    return len(rows)


def read_sheet(path, name):
    """Doc 1 sheet thanh list-of-lists (gia tri thuan) bang openpyxl read-only."""
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True, data_only=True)[name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # cat cac cot/dong rong o cuoi
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    w = max((len(r) for r in rows), default=0)
    while w and all(len(r) < w or r[w - 1] is None for r in rows):
        w -= 1
    return [r[:w] for r in rows]


if __name__ == '__main__':
    a = sys.argv[1:]
    if len(a) < 5 or a[2] != '--from':
        print(__doc__); sys.exit(2)
    dst, new_name, _, src, src_name = a[:5]
    rows = read_sheet(src, src_name)
    if '--add-cols' in a:
        spec = a[a.index('--add-cols') + 1]
        for item in spec.split(','):
            k, v = item.split('=', 1)
            rows[0].append(k)
            for r in rows[1:]:
                r.extend([None] * (len(rows[0]) - 1 - len(r)))
                try: r.append(int(v))
                except ValueError: r.append(v)
    n = add_sheet(dst, new_name, rows)
    print('%s: them sheet "%s" (%d dong, %d cot) tu %s::%s' % (dst, new_name, n, len(rows[0]) if rows else 0, src, src_name))
